from openpyxl import Workbook
from openpyxl import load_workbook
import os
from fpdf import FPDF
import smtplib
from email.message import EmailMessage
from dotenv import load_dotenv

#CREATING TXT

def extract_data_from_excel():
    wb = load_workbook('clients.xlsx')
    ws = wb.active

    if not os.path.exists("invoices"):
        os.mkdir("invoices")
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, email, item, qty, price, due = row
            total = qty * price
            invoice_text = f"""
            Invoice for {name}
            -------------------------
            Email: {email}
            Item: {item}
            Qty: {qty}
            Due date: {due}
            Price per item: ${price}
            Total: ${total}
            --------------------------
            Thanks for doing Business!
            """
            filename = f"invoices/{name.replace(' ', '_')}_invoice.txt"
            with open(filename, "w") as f:
                f.write(invoice_text)
        print("All invoices generated and saved.")
    except(AttributeError):
        print("Excel file is empty.")


#PDF CREATING
def convert_txt_to_pdf():
    invoice_records = []
    invoice_folder = "invoices"
    pdf_folder = "PDFs"
    os.makedirs(pdf_folder, exist_ok=True)

    # Loop through each invoice file
    for filename in os.listdir(invoice_folder):
        if filename.endswith(".txt"):
            filepath = os.path.join(invoice_folder, filename)

            # Read file
            with open(filepath, "r") as file:
                lines = []
                for line in file.readlines():
                    stripped = line.strip()
                    if stripped:
                        lines.append(stripped)

            # Extract fields
            customer_name = lines[0].replace("Invoice for ", "")
            email = lines[2].split("Email:")[1].strip()
            item = lines[3].split("Item:")[1].strip()
            qty = lines[4].split("Qty:")[1].strip()
            due_date = lines[5].split("Due date:")[1].strip()
            price = lines[6].split("Price per item:")[1].strip()
            total = lines[7].split("Total:")[1].strip()

            # Create PDF
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", size=14)
            pdf.cell(0, 10, f"Invoice for {customer_name}", ln=True)

            pdf.set_font("Helvetica", size=12)
            pdf.cell(0, 10, f"Email: {email}", ln=True)
            pdf.cell(0, 10, f"Item: {item}", ln=True)
            pdf.cell(0, 10, f"Quantity: {qty}", ln=True)
            pdf.cell(0, 10, f"Due Date: {due_date}", ln=True)
            pdf.cell(0, 10, f"Price per item: {price}", ln=True)
            pdf.set_font("Helvetica", style="B", size=12)
            pdf.cell(0, 10, f"Total: {total}", ln=True)

            pdf.ln(10)
            pdf.set_font("Helvetica", size=12)
            pdf.cell(0, 10, "Thanks for doing Business!", ln=True)

            # Save PDF
            name_part = filename.split("_")[0]
            pdf_file = f"{name_part}_invoice.pdf"
            pdf_path = os.path.join(pdf_folder, pdf_file)
            pdf.output(pdf_path)

            print(f"✅ Generated PDF for {customer_name}")
            invoice_records.append((email, pdf_file))
    return invoice_records    
#SENDING EMAIL

def send_email(cust_data):
    load_dotenv()
    user_email = input("Enter your email: ")
    user_app_pass = input("Enter your app password: ")

    smtp = smtplib.SMTP_SSL("smtp.gmail.com", 465)
    smtp.login(user_email, user_app_pass)

    for email, filename in cust_data:
        pdf_path = os.path.join("PDFs", filename)
        with open(pdf_path, "rb") as f:
            pdf_data = f.read()
        customer = filename.replace("_invoice.pdf", "").replace("_", " ")
        msg = EmailMessage()
        msg['From'] = user_email
        msg['To'] = email
        msg['Subject'] = f"Invoice for {customer}"
        msg.set_content(f"Dear {filename.split('_')[0]},\n\nPlease find your attached invoice.\n\nThanks!")

        msg.add_attachment(pdf_data, maintype="application", subtype="pdf", filename=filename)

        try:
            smtp.send_message(msg)
            print(f"📧 Invoice sent to {email}")
        except Exception as e:
            print(f"❌ Failed to send to {email}: {e}")

if __name__ == "__main__":
    extract_data_from_excel()
    customer_data = convert_txt_to_pdf()
    send_email(customer_data)